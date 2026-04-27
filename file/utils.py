from langchain_community.document_loaders import Docx2txtLoader, PyPDFLoader, CSVLoader, SitemapLoader, YoutubeLoader, \
    UnstructuredURLLoader, SeleniumURLLoader
from selenium.webdriver.support.ui import WebDriverWait
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from youtube_transcript_api import YouTubeTranscriptApi, TranscriptsDisabled, NoTranscriptFound, VideoUnavailable
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
from bs4 import BeautifulSoup
import time
from selenium.webdriver.support import expected_conditions as EC
from urllib.parse import urljoin, urlparse
import xml.etree.ElementTree as ET
from gtts import gTTS
import os
import openpyxl
import re
from concurrent.futures import ThreadPoolExecutor, as_completed
import threading
import queue
from concurrent.futures import ThreadPoolExecutor, as_completed
import threading


def _documents_to_text(docs):
    return "\n\n".join(
        doc.page_content for doc in docs
        if getattr(doc, 'page_content', None)
    ).strip()


def load_docx_file(file_path):
    loader = Docx2txtLoader(file_path)
    return _documents_to_text(loader.load())


def load_pdf_file(file_path):
    loader = PyPDFLoader(file_path)
    return _documents_to_text(loader.load())


def load_csv_file(file_path):
    loader = CSVLoader(file_path)
    return _documents_to_text(loader.load())


def load_xlsx_file(file_path):
    try:
        # Load the workbook
        workbook = openpyxl.load_workbook(file_path)

        # Initialize a list to store the content
        all_content = []

        # Iterate through all sheets
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]

            # Read all rows and columns in the sheet
            for row in sheet.iter_rows(values_only=True):
                # Join each cell in the row into a single string and add to the list
                row_content = ' '.join([str(cell) for cell in row if cell is not None])
                all_content.append(row_content)

        # Join all rows together into one large string
        combined_content = '\n'.join(all_content)

        return combined_content

    except FileNotFoundError:
        print(f"File not found: {file_path}")
        return None
    except Exception as e:
        print(f"An error occurred: {e}")
        return None


def load_sitemap_file(sitemap_url):
    loader = SitemapLoader(web_path=sitemap_url)
    docs = loader.load()
    if not docs:
        raise ValueError("No content could be extracted from the sitemap.")
    # Extract page_content from Langchain Document objects to form a single string
    content = "\n".join([doc.page_content for doc in docs if hasattr(doc, 'page_content')])
    if not content.strip():
        raise ValueError("Sitemap content is empty after extraction.")
    return content


def load_youtube_file(url):
    try:
        # Handle empty or non-string input
        if not url or not isinstance(url, str):
            raise ValueError("URL cannot be empty or must be a string.")

        # Clean the URL
        url = url.strip()

        # Extract video ID using a more robust regex or comprehensive parsing
        import re
        video_id_match = re.search(r'(?:v=|\/)([0-9A-Za-z_-]{11}).*', url)
        if video_id_match:
            video_id = video_id_match.group(1)
        else:
            raise ValueError("Could not extract a valid 11-character YouTube video ID.")

        # Fetch the transcript using the instance method confirmed to work in this environment
        api = YouTubeTranscriptApi()
        fetched_transcript = api.fetch(video_id)

        if not fetched_transcript:
            raise ValueError("No transcript content available.")

        full_text = []
        for snippet in fetched_transcript:
            # Handle both dictionary and object formats for snippet
            if isinstance(snippet, dict):
                full_text.append(snippet.get('text', ''))
            else:
                full_text.append(getattr(snippet, 'text', ''))

        content = ' '.join(full_text)
        if not content.strip():
            raise ValueError("Fetched transcript is empty.")

        return content
    except TranscriptsDisabled:
        raise ValueError("Transcripts are disabled for this video.")
    except NoTranscriptFound:
        raise ValueError("No transcript found for this video.")
    except VideoUnavailable:
        raise ValueError("The video is unavailable.")
    except Exception as e:
        error_msg = str(e)
        if len(error_msg) > 200:
            error_msg = error_msg[:197] + "..."
        raise ValueError(f"Could not load youtube video: {error_msg}")


def load_multiple_url(urls):
    url = [urls]
    loader = UnstructuredURLLoader(url, ssl_verify=False)
    return loader.load()


def load_url(urls):
    url = [urls]

    loader = SeleniumURLLoader(urls=url)

    return loader.load()


def loads_urls(urls):
    options = Options()
    options.add_argument("--headless=new")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--disable-gpu")
    options.add_argument("--disable-notifications")
    options.add_argument("--remote-debugging-port=0")
    options.add_argument("start-maximized")
    options.add_argument("--window-size=1920,1080")
    options.add_argument("enable-automation")
    options.add_argument("--disable-infobars")
    options.add_argument("--disable-browser-side-navigation")
    options.add_argument("--disable-features=VizDisplayCompositor")
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) "
                         "Chrome/120.0.0.0 Safari/537.36")

    service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service, options=options)

    # Limit to 25 URLs
    urls = urls[:25]

    all_text_content = []
    try:
        for url in urls:
            try:
                driver.get(url)
                time.sleep(5)  # Initial wait for static content

                # Scroll to bottom to trigger lazy-loaded content
                last_height = driver.execute_script("return document.body.scrollHeight")
                for _ in range(5):  # Limit scrolls to prevent infinite loops on some sites
                    driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                    time.sleep(2)
                    new_height = driver.execute_script("return document.body.scrollHeight")
                    if new_height == last_height:
                        break
                    last_height = new_height

                # Wait for any Final dynamic elements
                WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.TAG_NAME, 'body')))

                soup = BeautifulSoup(driver.page_source, 'html.parser')

                # Remove script and style elements to get cleaner content
                for element in soup(["script", "style"]):
                    element.decompose()

                # Get clean, separated text
                text_content = soup.get_text(separator='\n', strip=True)
                if text_content:
                    all_text_content.append(f"--- URL: {url} ---\n{text_content}")
            except Exception as e:
                print(f"Error fetching URL {url}: {e}")

        # Join all text contents together
        full_text_content = "\n\n".join(all_text_content)
        if not full_text_content.strip():
            return None
        return full_text_content
    except Exception as e:
        print(f"An unexpected error occurred: {e}")
        return None
    finally:
        driver.quit()


def get_base_url(url):
    return url


def generate_sitemap(base_url):
    """Generates a sitemap for the given base URL using a multi-threaded crawler."""
    if not base_url:
        raise ValueError("base_url cannot be None or empty")
    if not base_url.startswith(('http://', 'https://')):
        raise ValueError("base_url must start with 'http://' or 'https://'")

    visited_urls = set()
    url_queue = queue.Queue()
    url_queue.put(base_url)
    visited_urls.add(base_url)

    lock = threading.Lock()

    def crawl_worker():
        options = Options()
        options.add_argument("--headless=new")
        options.add_argument("--no-sandbox")
        options.add_argument("--disable-dev-shm-usage")
        options.add_argument("--disable-gpu")
        options.add_argument("--disable-notifications")
        options.add_argument("--window-size=1920,1080")
        options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36")

        driver = None
        try:
            service = Service(ChromeDriverManager().install())
            driver = webdriver.Chrome(service=service, options=options)

            while True:
                try:
                    # Wait for a URL with a timeout to allow exit
                    current_url = url_queue.get(timeout=10)
                except queue.Empty:
                    break

                try:
                    driver.get(current_url)
                    time.sleep(2)
                    soup = BeautifulSoup(driver.page_source, 'html.parser')

                    for link in soup.find_all("a", href=True):
                        link_url = urljoin(current_url, link["href"])
                        parsed_url = urlparse(link_url)
                        normalized_url = parsed_url._replace(fragment="").geturl()

                        if normalized_url.startswith(base_url):
                            with lock:
                                if normalized_url not in visited_urls and len(visited_urls) < 25:
                                    visited_urls.add(normalized_url)
                                    url_queue.put(normalized_url)
                except Exception as e:
                    print(f"Error crawling {current_url}: {e}")
                finally:
                    url_queue.task_done()
        finally:
            if driver:
                driver.quit()

    # Start 3 worker threads for crawling
    # Crawling is recursive, so we use a pool
    threads = []
    for _ in range(3):
        t = threading.Thread(target=crawl_worker)
        t.start()
        threads.append(t)

    # Wait for all tasks to be completed
    url_queue.join()

    # Wait for threads to finish
    for t in threads:
        t.join()

    # Save the sitemap to the user's desktop
    desktop_path = os.path.join(os.path.expanduser("~"), "Desktop", "sitemap.txt")
    os.makedirs(os.path.dirname(desktop_path), exist_ok=True)
    with open(desktop_path, "w", encoding="utf-8") as f:
        for url in sorted(visited_urls):
            f.write(url + "\n")

    return desktop_path


def extract_and_save_content(sitemap_xml_content, proj_id):
    """
    Extracts URLs from a sitemap XML content, fetches content in parallel, and saves it.
    """
    from .models import Project  # Local import to avoid circular dependencies

    def extract_urls_from_sitemap(xml_content):
        root = ET.fromstring(xml_content)
        urls = []
        for url in root.findall('{http://www.sitemaps.org/schemas/sitemap/0.9}url'):
            loc = url.find('{http://www.sitemaps.org/schemas/sitemap/0.9}loc').text
            urls.append(loc)
        return urls[:25] # Limit to 25 URLs

    def process_url(url):
        """Worker function to process a single URL."""
        options = Options()
        options.add_argument("--headless=new")
        options.add_argument("--no-sandbox")
        options.add_argument("--disable-dev-shm-usage")
        options.add_argument("--disable-gpu")
        options.add_argument("--disable-notifications")
        options.add_argument("--window-size=1920,1080")
        options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36")

        driver = None
        try:
            service = Service(ChromeDriverManager().install())
            driver = webdriver.Chrome(service=service, options=options)
            driver.get(url)
            time.sleep(3)  # Reduced wait time for parallel workers

            # Scroll to bottom
            last_height = driver.execute_script("return document.body.scrollHeight")
            for _ in range(3):
                driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                time.sleep(1.5)
                new_height = driver.execute_script("return document.body.scrollHeight")
                if new_height == last_height:
                    break
                last_height = new_height

            WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.TAG_NAME, "body")))

            soup = BeautifulSoup(driver.page_source, 'html.parser')
            for element in soup(["script", "style"]):
                element.decompose()

            text = soup.get_text(separator='\n', strip=True)
            return f"--- URL: {url} ---\n{text}" if text else None
        except Exception as e:
            print(f"Error processing URL {url}: {e}")
            return None
        finally:
            if driver:
                driver.quit()

    urls = extract_urls_from_sitemap(sitemap_xml_content)
    all_contents = []

    # Use ThreadPoolExecutor for parallel extraction
    # Limiting to 5 workers to prevent heavy resource usage
    with ThreadPoolExecutor(max_workers=5) as executor:
        future_to_url = {executor.submit(process_url, url): url for url in urls}
        for future in as_completed(future_to_url):
            content = future.result()
            if content:
                all_contents.append(content)

    # Single database update at the end for thread safety and efficiency
    if all_contents:
        try:
            project = Project.objects.get(pk=proj_id)
            combined_text = "\n\n".join(all_contents)
            project.content = (project.content + "\n\n" + combined_text) if project.content else combined_text
            project.save()
        except Exception as e:
            print(f"Error saving project content: {e}")


def text_to_speech(text):
    # Determine the desktop directory path
    desktop_dir = os.path.join(os.path.expanduser("~"), "Desktop")

    # Ensure the desktop directory exists (it should, by default)
    if not os.path.exists(desktop_dir):
        raise FileNotFoundError("Desktop directory not found")

    # Create a file name and full path for the audio file
    audio_file = f"{text[:10].replace(' ', '_')}.mp3"
    audio_file_path = os.path.join(desktop_dir, audio_file)

    # Convert text to speech and save it to the desktop
    tts = gTTS(text=text, lang='en')
    tts.save(audio_file_path)

    return audio_file_path
